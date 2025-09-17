import openpyxl
import re
from datasets import load_dataset
from pathlib import Path
import json

TASKS_NAME = ["随访信息来源", "受试者是否死亡", "受试者是否住院", "受试者是否手术", "受试者是否用药"]

def load_jsonl(pth):
    dataset = load_dataset("json", data_files=str(pth), split="train")
    return [i for i in dataset]

def parse_data(data_string):
    lines = data_string.strip().split('\n')
    try:
        # load by json
        data = json.loads(data_string)
        return data
    except json.decoder.JSONDecodeError:
        pass
    data = {}
    valid_values = set(["是", "否", "不确定", "未提及", "本人", "亲属"])

    for line in lines:
        if "：" not in line:
            continue
        key, value = line.split('：', 1)
        key = key.strip()
        
        if key == "受试者是否用药":
            value = re.search(r'^(是|否|不确定|未提及)', value)
            if value:
                data[key] = value.group(1)
            else:
                data[key] = "未提及"
        else:
            data[key] = value.strip()

    return data

def generic_export(input_data, output_file, data_processor):
    sorted_data = sorted(input_data, key=lambda x: x.get('dCap', ''))
    
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = 'dCap'
    for idx, task in enumerate(TASKS_NAME, start=2):
        ws.cell(row=1, column=idx, value=f"Pred_{task}")

    for row, item in enumerate(sorted_data, start=2):
        ws.cell(row=row, column=1, value=item.get('dCap', ''))
        processed_item = data_processor(item)
        for idx, task in enumerate(TASKS_NAME, start=2):
            ws.cell(row=row, column=idx, value=processed_item.get(task, ''))

    wb.save(output_file)


def export_finetune_qwen7b(data_identifier):
    output_file = Path(f"data/output/finetune_qwen2_7b_{data_identifier}/follow_up_train_data_{data_identifier}.xlsx")
    output_file.parent.mkdir(exist_ok=True)

    trained_model_dir = Path("data/trained_models")
    ref_dir = Path(f"data/follow_up_train_data_{data_identifier}")
    data = []
    for data_pth in trained_model_dir.glob(f"qwen2_7b_instruct_lora_sft_{data_identifier}_f*_v0/predict/generated_predictions.jsonl"):
        data += load_jsonl(data_pth)

    ref_data = []
    for ref_pth in ref_dir.glob("test_fold_*.jsonl"):
        ref_data += load_jsonl(ref_pth)
    
    mapping = {i["input"]: i["dCap"] for i in ref_data}
    
    for item in data:
        input_split = item["prompt"].split("\nassistant\n")[0].split("导致无法做出判断\n")[-1]
        item["dCap"] = mapping[input_split]
    
    generic_export(data, output_file, lambda item: parse_data(item.get("predict")))


if __name__ == "__main__":
    export_finetune_qwen7b("aug")
    export_finetune_qwen7b("wo_aug")
